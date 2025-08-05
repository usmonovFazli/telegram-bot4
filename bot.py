import logging
import os
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, InputFile
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ChatMemberHandler,
    ContextTypes,
    filters,
)

from database import (
    init_db,
    add_or_update_channel,
    update_channel_status,
    get_channels,
    increment_video_count,
)

# Загрузка переменных окружения
load_dotenv()
TOKEN = os.getenv("BOT_TOKEN", "")
AUTHORIZED_PASSWORD = os.getenv("BOT_PASSWORD", "")
LEAVE_PASSWORD = "1234"

authorized_users = set()
leave_confirmations = {}

# Главное меню
MAIN_MENU = ReplyKeyboardMarkup(
    [["🎥 Отправить видео", "📊 Статистика"], ["📥 Экспорт Excel", "🚪 Покинуть чаты"]],
    resize_keyboard=True
)

logging.basicConfig(level=logging.INFO)

# --- Авторизация ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id in authorized_users:
        await update.message.reply_text("✅ Добро пожаловать снова!", reply_markup=MAIN_MENU)
    else:
        await update.message.reply_text("🔐 Введите пароль для доступа:")
        context.user_data["awaiting_password"] = True

async def handle_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text = update.message.text.strip()

    if context.user_data.get("awaiting_password"):
        if text == AUTHORIZED_PASSWORD:
            authorized_users.add(user_id)
            context.user_data["awaiting_password"] = False
            await update.message.reply_text("✅ Доступ разрешён!", reply_markup=MAIN_MENU)
        else:
            await update.message.reply_text("❌ Неверный пароль. Попробуйте снова:")
    elif leave_confirmations.get(user_id) == "password":
        await handle_leave_password(update, context)

# --- Проверка доступа ---
def check_access(user_id):
    return user_id in authorized_users

# --- Обработка текста ---
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await handle_password(update, context)

# --- Видео ---
async def prompt_video(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not check_access(update.effective_user.id):
        await update.message.reply_text("⛔️ У вас нет доступа. Введите /start и пароль.")
        return
    await update.message.reply_text("📤 Пожалуйста, отправьте видео, которое хотите разослать.")

async def handle_video(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not check_access(update.effective_user.id):
        await update.message.reply_text("⛔️ У вас нет доступа.")
        return

    if not update.message or not update.message.video:
        return

    video = update.message.video
    caption = update.message.caption or "📹 Новое видео:"
    chats = get_channels()

    count = 0
    for chat_id, *_ in chats:
        try:
            await context.bot.send_video(chat_id=chat_id, video=video.file_id, caption=caption)
            increment_video_count(chat_id)
            count += 1
        except Exception as e:
            logging.warning(f"❌ Не удалось отправить в {chat_id}: {e}")

    await update.message.reply_text(f"✅ Видео отправлено в {count} чатов.")

# --- Обработка входа/выхода из чатов ---
async def chat_member_update(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.my_chat_member.chat
    new_status = update.my_chat_member.new_chat_member.status

    try:
        members = await context.bot.get_chat_member_count(chat.id)
    except Exception as e:
        members = -1
        logging.warning(f"Не удалось получить количество участников: {e}")

    if members != -1 and members < 1:
        try:
            update_channel_status(chat.id, chat_type="left")  # ✅ ДО выхода
            await context.bot.leave_chat(chat.id)
            logging.info(f"🚪 Покинул чат {chat.title} — участников меньше 50 ({members})")
        except Exception as e:
            logging.warning(f"❌ Не удалось выйти из {chat.title}: {e}")
        return


    link = f"https://t.me/{chat.username}" if chat.username else ""
    title = chat.title or "Без названия"

    try:
        add_or_update_channel(chat.id, title, members, new_status, link)
        logging.info(f"✅ Чат зарегистрирован или обновлён: {title} ({new_status}) — {chat.id}")
    except Exception as e:
        logging.error(f"❌ Ошибка при добавлении/обновлении чата: {e}")

# --- Обновление данных чатов ---
async def refresh_members(context: ContextTypes.DEFAULT_TYPE):
    chats = get_channels()
    for chat_id, title, _, _, _, _, link in chats:
        try:
            members = await context.bot.get_chat_member_count(chat_id)
            chat = await context.bot.get_chat(chat_id)
            update_channel_status(chat_id, title=chat.title, members=members, chat_type=chat.type, link=link)
        except Exception as e:
            logging.warning(f"⚠️ Не удалось обновить {chat_id}: {e}")
            update_channel_status(chat_id, chat_type="left")

# --- Статистика ---
async def show_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not check_access(update.effective_user.id):
        await update.message.reply_text("⛔️ У вас нет доступа.")
        return

    await update.message.reply_text("♻️ Обновляю информацию о каналах...")
    await refresh_members(context)

    chats = get_channels()
    if not chats:
        await update.message.reply_text("⚠️ Нет подключённых каналов/групп.")
        return

    stats = {
        "channel": {"active": 0, "inactive": 0},
        "group": {"active": 0, "inactive": 0},
        "supergroup": {"active": 0, "inactive": 0},
    }

    for _, _, _, _, _, chat_type, _ in chats:
        if chat_type in ["left", "kicked"]:
            for t in ["channel", "group", "supergroup"]:
                stats[t]["inactive"] += 0  # просто чтобы пропустить
            continue

        if chat_type in stats:
            stats[chat_type]["active"] += 1
        elif chat_type in ["left", "kicked"]:
            pass  # отдельная обработка ниже

    # Обработка удалённых отдельно
    for _, _, _, _, _, chat_type, _ in chats:
        if chat_type in ["left", "kicked"]:
            # Мы не знаем, что это — канал или группа, но можно посчитать как "unknown" или оставить
            continue
        if chat_type in stats:
            continue  # уже учтён
    # Теперь пересчитаем активные и неактивные по каждой категории
    for chat in chats:
        _, _, _, _, _, chat_type, _ = chat
        if chat_type in stats:
            if chat_type in ["channel", "group", "supergroup"]:
                if chat_type in stats:
                    if chat_type in ["left", "kicked"]:
                        stats[chat_type]["inactive"] += 1
                else:
                    stats[chat_type] = {"active": 0, "inactive": 1}
        elif chat_type in ["left", "kicked"]:
            pass

    # Собираем текст
    total = len(chats)
    text = "📊 Общая статистика:\n\n"
    for t, label in [("channel", "Каналы"), ("group", "Группы"), ("supergroup", "Супергруппы")]:
        active = stats[t]["active"]
        inactive = stats[t]["inactive"]
        total_type = active + inactive
        text += f"• {label}: {total_type} ({active} активных / {inactive} удалённых)\n"

    text += f"\nВсего чатов: {total}"

    await update.message.reply_text(text)


# --- Экспорт Excel ---
async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not check_access(update.effective_user.id):
        await update.message.reply_text("⛔️ У вас нет доступа.")
        return

    await update.message.reply_text("📦 Обновляю данные перед экспортом...")
    await refresh_members(context)

    chats = get_channels()
    if not chats:
        await update.message.reply_text("⚠️ Нет данных для экспорта.")
        return

    active = [c for c in chats if c[5] not in ["left", "kicked"]]
    left = [c for c in chats if c[5] in ["left", "kicked"]]
    sorted_chats = active + left

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каналы и группы"
    ws.append(["ID", "Название", "Участники", "Отправлено видео", "Дата добавления", "Тип", "Ссылка"])

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for row in sorted_chats:
        id_, title, members, videos, date_added, chat_type, link = row
        date_str = date_added.strftime('%Y-%m-%d %H:%M') if isinstance(date_added, datetime) else str(date_added)
        data = [id_, title, members, videos, date_str, chat_type, link]
        ws.append(data)
        fill = red_fill if chat_type in ["left", "kicked"] else green_fill
        for col in range(1, len(data) + 1):
            ws.cell(row=ws.max_row, column=col).fill = fill

    file_path = "channels_export.xlsx"
    wb.save(file_path)
    with open(file_path, "rb") as f:
        await update.message.reply_document(InputFile(f, filename="channels_export.xlsx"))
    os.remove(file_path)

# --- Покинуть все чаты ---
async def initiate_leave(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not check_access(user_id):
        await update.message.reply_text("⛔️ У вас нет доступа.")
        return

    leave_confirmations[user_id] = "confirm"
    markup = ReplyKeyboardMarkup([["✅ Да", "❌ Нет"]], resize_keyboard=True, one_time_keyboard=True)
    await update.message.reply_text("⚠️ Вы уверены, что хотите выйти из всех чатов?", reply_markup=markup)

async def handle_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text = update.message.text.strip()

    if leave_confirmations.get(user_id) != "confirm":
        return

    if text == "✅ Да":
        leave_confirmations[user_id] = "password"
        await update.message.reply_text("🔐 Введите пароль для подтверждения:")
    else:
        leave_confirmations.pop(user_id, None)
        await update.message.reply_text("❎ Операция отменена.", reply_markup=MAIN_MENU)

async def handle_leave_password(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text = update.message.text.strip()

    if leave_confirmations.get(user_id) != "password":
        return

    if text == LEAVE_PASSWORD:
        chats = get_channels()  # ✅ ЭТОЙ СТРОКИ НЕ ХВАТАЛО
        left = 0
        for chat_id, *_ in chats:
            try:
                update_channel_status(chat_id, chat_type="left")  # ✅ фикс статуса
                await context.bot.leave_chat(chat_id)
                left += 1
            except Exception as e:
                logging.warning(f"❌ Не смог выйти из {chat_id}: {e}")

        leave_confirmations.pop(user_id, None)
        await update.message.reply_text(f"🚪 Вышел из {left} чатов.", reply_markup=MAIN_MENU)
    else:
        await update.message.reply_text("❌ Неверный пароль. Операция отменена.")
        leave_confirmations.pop(user_id, None)

# --- Запуск ---
def main():
    init_db()
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(ChatMemberHandler(chat_member_update, ChatMemberHandler.MY_CHAT_MEMBER))
    app.add_handler(MessageHandler(filters.Regex("^🎥 Отправить видео$"), prompt_video))
    app.add_handler(MessageHandler(filters.Regex("^📊 Статистика$"), show_stats))
    app.add_handler(MessageHandler(filters.Regex("^📥 Экспорт Excel$"), export_excel))
    app.add_handler(MessageHandler(filters.Regex("^🚪 Покинуть чаты$"), initiate_leave))
    app.add_handler(MessageHandler(filters.Regex("^(✅ Да|❌ Нет)$"), handle_confirmation))
    app.add_handler(MessageHandler(filters.VIDEO & filters.ChatType.PRIVATE, handle_video))
    app.add_handler(MessageHandler(filters.TEXT & filters.ChatType.PRIVATE, handle_text))

    logging.info("Бот запущен...")
    app.run_polling()

if __name__ == "__main__":
    main()
